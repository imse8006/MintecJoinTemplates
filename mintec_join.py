import pandas as pd
import os
import warnings
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# ignore errors
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

categories = {
    "Catégorie 1": ["AddedCosts06 - Packaging UK Data Export", "AddedCosts07 - Packaging EU Data Export", "AddedCosts08 - Packaging US Data Export", "AddedCosts01 - Distribution UK Data Export", "Added Costs - 2 - Distribution EU Data Export", "AddedCosts03 - Distribution - World Data Export"],
    "Catégorie 2": ["PACKAGING8", "PACKAGING5", "PACKAGING3", "PACKAGING4"],
    "Catégorie 3": ["Electricity base month ahead UK Data Export", "Electricity base EEX FR Data Export", "Electricity ind US Data Export", "Natural gas Eur Data Export", "Natural gas NBPI UK Data Export", "Crude oil Brent ICE EU Data Export", "Crude oil Brent fob UK Data Export", "Crude oil ave spot Data Export"]
}

category_colors = {"Catégorie 1": "FFD966", "Catégorie 2": "A9D08E", "Catégorie 3": "9BC2E6"}

source_folder = Path(r"C:\Users\IL00030293\OneDrive - Sysco Corporation\Documents\PGM\Mintec downloads\January\files")
output_file = source_folder / "Merged_Data.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Merged Data"

col_idx = 1

for file in categories["Catégorie 1"] + categories["Catégorie 2"] + categories["Catégorie 3"]:
    file_path = source_folder / f"{file}.xlsx"

    if not file_path.exists():
        print(f"⚠️ Fichier introuvable : {file_path}")
        continue

    try:
        df = pd.read_excel(file_path, usecols="B", skiprows=1)  # Ajusté pour inclure B3
        column_data = df.iloc[:, 0].dropna().tolist()

        if not column_data:
            print(f"⚠️ Aucune donnée trouvée dans {file}.xlsx (colonne B). Vérifie que B3 contient bien des valeurs.")
            continue

        ws.cell(row=1, column=col_idx, value=file)

        for row_idx, value in enumerate(column_data, start=2):
            ws.cell(row=row_idx, column=col_idx, value=value)

        for category, file_list in categories.items():
            if file in file_list:
                color = category_colors[category]
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                ws.cell(row=1, column=col_idx).fill = fill

        col_idx += 1

    except Exception as e:
        print(f"❌ Erreur avec {file}.xlsx : {e}")

try:
    wb.save(output_file)
    print(f"✅ Fusion terminée ! Fichier sauvegardé sous : {output_file}")
except PermissionError:
    print(f"❌ Erreur : Impossible d'écrire dans {output_file}. Fermez-le dans Excel et réessayez.")
